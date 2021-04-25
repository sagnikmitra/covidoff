from openpyxl import load_workbook
import openpyxl as pxl
import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from bokeh.models.widgets import Button
from bokeh.models import CustomJS
from streamlit_bokeh_events import streamlit_bokeh_events
from io import StringIO


gc = pd.read_excel('gc.xlsx')
exp = pxl.load_workbook('gc.xlsx')
sheet = exp['Sister Nivedita University, Kol']
st.write("""
# Check your Qwiklabs Progress for GoogleCloudFacilitator Program 2021
## Your Host: Sister Nivedita University
""")
flag = 0
milestone1_quest = 8
milestone1_skill_badges = 4
milestone2_quest = 16
milestone2_skill_badges = 8
milestone3_quest = 24
milestone3_skill_badges = 12
milestone4_quest = 30
milestone4_skill_badges = 15

# copy_button = Button(label="Get Clipboard Data")
# copy_button.js_on_event("button_click", CustomJS(code="""
#     navigator.clipboard.readText().then(text => document.dispatchEvent(new CustomEvent("GET_TEXT", {detail: text})))
#     """))
# result = streamlit_bokeh_events(
#     copy_button,
#     events="GET_TEXT",
#     key="get_text",
#     refresh_on_update=False,
#     override_height=75,
#     debounce_time=0)

# if result:
#     if "GET_TEXT" in result:
#         df = pd.read_csv(StringIO(result.get("GET_TEXT")))
#         st.table(df)

selected = st.sidebar.selectbox(
    "Search By", ("Default", "Email", "Public Profile URL"))

page_name = ['Email', 'Public Profile URL']
page = st.radio('Search By', page_name)

# if page == 'Name':
#     str = st.text_input('Enter You Qwiklabs Name')
#     for i in range(2, sheet.max_row+1):
#         if sheet.cell(row=i, column=1).value.lower() == str.lower():
#             st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
#             st.write(f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
#             st.write(f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
#             st.write("## **Qwiklabs Public URL:**")
#             st.write(sheet.cell(row=i, column=6).value)
#             st.write(f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
#             st.write(f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")
#             quest_difference_milestone_1 = (milestone1_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_1 = (milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1) :
#                 st.write("## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
#             else:
#                 st.warning("You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
#                 st.write(f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
#             quest_difference_milestone_2 = (milestone2_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_2 = (milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1) :
#                 st.write("## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
#             quest_difference_milestone_3 = (milestone3_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_3 = (milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1) :
#                 st.write("## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
#             quest_difference_milestone_4 = (milestone4_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_4 = (milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1) :
#                 st.write("## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_4} Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
#             flag = flag + 1

#     if flag == 0:
#         st.warning("No Search Found")

if page == 'Email':
    str = st.text_input('Enter You Qwiklabs Email Id')
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=2).value.lower() == str.lower():
            st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
            st.write(
                f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
            st.write(
                f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
            st.write("## **Qwiklabs Public URL:**")
            st.write(sheet.cell(row=i, column=6).value)
            st.write(
                f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
            st.write(
                f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")

            #Quest and Skill Badges difference calculation for Milestone 1
            quest_difference_milestone_1 = (
                milestone1_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_1 = (
                milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))

            #Milestone 1
            if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1):
                st.write(
                    "## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
            else:
                st.warning("You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
                if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                else:
                    if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                    if (skill_badges_difference_milestone_1 <= 0 and quest_difference_milestone_1 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_1} Quests** and **0 Skill Badges** Away to Complete the First Milestone")

            #Quest and Skill Badges difference calculation for Milestone 2
            quest_difference_milestone_2 = (
                milestone2_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_2 = (
                milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))

            #Milestone 2
            if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
            else:
                if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                else:
                    if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                    if (skill_badges_difference_milestone_2 <= 0 and quest_difference_milestone_2 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_2} Quests** and **0 Skill Badges** Away to Complete the Second Milestone")

            #Quest and Skill Badges difference calculation for Milestone 3
            quest_difference_milestone_3 = (
                milestone3_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_3 = (
                milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))

            #Milestone 3
            if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
            else:
                if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                else:
                    if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                    if (skill_badges_difference_milestone_3 <= 0 and quest_difference_milestone_3 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_3} Quests** and **0 Skill Badges** Away to Complete the Third Milestone")

            #Quest and Skill Badges difference calculation for Milestone 4
            quest_difference_milestone_4 = (
                milestone4_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_4 = (
                milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))

            #Milestone 4
            if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
            else:
                if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_4} Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                else:
                    if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                    if (skill_badges_difference_milestone_4 <= 0 and quest_difference_milestone_4 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_4} Quests** and **0 Skill Badges** Away to Complete the Ultimate Milestone")

            flag = flag + 1

    if flag == 0:
        st.write("No Search Found for the entered Details")

if page == 'Public Profile URL':
    str = st.text_input('Enter You Qwiklabs Public URL')
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=6).value.lower() == str.lower():
            st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
            st.write(
                f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
            st.write(
                f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
            st.write("## **Qwiklabs Public URL:**")
            st.write(sheet.cell(row=i, column=6).value)
            st.write(
                f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
            st.write(
                f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")

            #Quest and Skill Badges difference calculation for Milestone 1
            quest_difference_milestone_1 = (
                milestone1_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_1 = (
                milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))

            #Milestone 1
            if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1):
                st.write(
                    "## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
            else:
                st.warning("You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
                if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                else:
                    if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                    if (skill_badges_difference_milestone_1 <= 0 and quest_difference_milestone_1 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_1} Quests** and **0 Skill Badges** Away to Complete the First Milestone")

            #Quest and Skill Badges difference calculation for Milestone 2
            quest_difference_milestone_2 = (
                milestone2_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_2 = (
                milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))

            #Milestone 2
            if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
            else:
                if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                else:
                    if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                    if (skill_badges_difference_milestone_2 <= 0 and quest_difference_milestone_2 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_2} Quests** and **0 Skill Badges** Away to Complete the Second Milestone")

            #Quest and Skill Badges difference calculation for Milestone 3
            quest_difference_milestone_3 = (
                milestone3_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_3 = (
                milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))

            #Milestone 3
            if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
            else:
                if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                else:
                    if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                    if (skill_badges_difference_milestone_3 <= 0 and quest_difference_milestone_3 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_3} Quests** and **0 Skill Badges** Away to Complete the Third Milestone")

            #Quest and Skill Badges difference calculation for Milestone 4
            quest_difference_milestone_4 = (
                milestone4_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_4 = (
                milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))

            #Milestone 4
            if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
            else:
                if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Ultimate Milestone")
                else:
                    if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                    if (skill_badges_difference_milestone_4 <= 0 and quest_difference_milestone_4 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_4} Quests** and **0 Skill Badges** Away to Complete the Ultimate Milestone")

            flag = flag + 1

    if flag == 0:
        st.write("No Search Found for the entered Details")

# if selected == "Name":
#     str = st.text_input('Enter You Qwiklabs Name')
#     for i in range(2, sheet.max_row+1):
#         if sheet.cell(row=i, column=1).value.lower() == str.lower():
#             st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
#             st.write(f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
#             st.write(f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
#             st.write("## **Qwiklabs Public URL:**")
#             st.write(sheet.cell(row=i, column=6).value)
#             st.write(f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
#             st.write(f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")
#             quest_difference_milestone_1 = (milestone1_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_1 = (milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1) :
#                 st.write("## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
#             else:
#                 st.warning("You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
#                 st.write(f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
#             quest_difference_milestone_2 = (milestone2_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_2 = (milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1) :
#                 st.write("## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
#             quest_difference_milestone_3 = (milestone3_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_3 = (milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1) :
#                 st.write("## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
#             quest_difference_milestone_4 = (milestone4_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_4 = (milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1) :
#                 st.write("## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_4} Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
#             flag = flag + 1

    # if flag == 0:
    #     st.write("No Search Found")

# if selected == "Email":
#     str = st.text_input('Enter You Qwiklabs Email Id')
#     for i in range(2, sheet.max_row+1):
#         if sheet.cell(row=i, column=2).value.lower() == str.lower():
#             st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
#             st.write(f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
#             st.write(f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
#             st.write("## **Qwiklabs Public URL:**")
#             st.write(sheet.cell(row=i, column=6).value)
#             st.write(f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
#             st.write(f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")
#             quest_difference_milestone_1 = (milestone1_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_1 = (milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1) :
#                 st.write("## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire: :star_struck:")
#             else:
#                 st.warning("You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
#                 st.write(f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
#             quest_difference_milestone_2 = (milestone2_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_2 = (milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1) :
#                 st.write("## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
#             quest_difference_milestone_3 = (milestone3_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_3 = (milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1) :
#                 st.write("## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
#             quest_difference_milestone_4 = (milestone4_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_4 = (milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1) :
#                 st.write("## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_4} Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
#             flag = flag + 1

#     if flag == 0:
#         st.write("No Search Found")

# if selected == "Public Profile URL":
#     str = st.text_input('Enter You Qwiklabs Public URL')
#     for i in range(2, sheet.max_row+1):
#         if sheet.cell(row=i, column=6).value.lower() == str.lower():
#             st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
#             st.write(f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
#             st.write(f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
#             st.write("## **Qwiklabs Public URL:**")
#             st.write(sheet.cell(row=i, column=6).value)
#             st.write(f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
#             st.write(f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")
#             quest_difference_milestone_1 = (milestone1_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_1 = (milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1) :
#                 st.write("## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
#             else:
#                 st.warning("You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
#                 st.write(f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
#             quest_difference_milestone_2 = (milestone2_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_2 = (milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1) :
#                 st.write("## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
#             quest_difference_milestone_3 = (milestone3_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_3 = (milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1) :
#                 st.write("## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
#             quest_difference_milestone_4 = (milestone4_quest - int(sheet.cell(row=i, column=7).value))
#             skill_badges_difference_milestone_4 = (milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))
#             if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1) :
#                 st.write("## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
#             else:
#                 st.write(f"### You are **{quest_difference_milestone_4} Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
#             flag = flag + 1

#     if flag == 0:
#         st.write("No Search Found")

st.write(" ")
st.write(" ")
st.write(" ")
st.write(" ")
st.write(" ")
st.write(
    "#### Developed & Maintained by: **[Sagnik Mitra](https://linkedin.com/in/sagnikmitra/) & [Manish Kumar Barnwal](https://linkedin.com/in/imanishbarnwal/)** with :heart:")

# dataset_name = st.sidebar.selectbox("Select Dataset",("Iris","Breast Cancer","Wine Dataset"))
# st.write(f"## Name of the Dataset: {dataset_name}")


# # def get_quests:
# #     pass

# def get_dataset(dataset_name):
#     if dataset_name == "Iris":
#         data = datasets.load_iris()
#     elif dataset_name == "Breast Cancer":
#         data = datasets.load_breast_cancer()
#     else:
#         data = datasets.load_wine()
#     X = data.data
#     y = data.target
#     return X,y

# X, y = get_dataset(dataset_name)
# st.write("## Shape of Dataset:",X.shape)
# st.write("## Number of classes:",len(np.unique(y)))

# def add_parameter_ui(classifier_name):
#     params = dict()
#     if classifier_name == "KNN":
#         K = st.sidebar.slider("K",1,15)
#         params["K"] = K
#     elif classifier_name == "SVM":
#         C = st.sidebar.slider("C",0.01,10.0)
#         params["C"] = C
#     elif classifier_name == "Random Forest":
#         max_depth = st.sidebar.slider("max_depth",2,15)
#         n_estimators = st.sidebar.slider("n_estimators",1,100)
#         params["max_depth"] = max_depth
#         params["n_estimators"] = n_estimators
#     return params

# params = add_parameter_ui(classifier_name)

# def get_classifier(classifier_name, params):
#     if classifier_name == "KNN":
#         classifier = KNeighborsClassifier(n_neighbors=params["K"])
#     elif classifier_name == "SVM":
#         classifier = SVC(C=params["C"])
#     elif classifier_name == "Random Forest":
#         classifier = RandomForestClassifier(n_estimators=params["n_estimators"],max_depth=params["max_depth"],random_state=1234)
#     return classifier

# classifier = get_classifier(classifier_name,params)

# X_train, X_test, y_train, y_test = train_test_split(X,y,test_size = 0.2, random_state = 1234)

# classifier.fit(X_train,y_train)
# y_pred = classifier.predict(X_test)

# accuracy_score = accuracy_score(y_test, y_pred)

# st.write(f"## Classifier = {classifier_name}")
# st.write(f"## Accuracy = {accuracy_score}")

# pca = PCA(2)
# X_projected = pca.fit_transform(X)

# x1 = X_projected[:,0]
# x2 = X_projected[:,1]

# fig = plt.figure()
# plt.scatter(x1,x2,c=y,alpha=0.8,cmap = "viridis")
# plt.xlabel("Principal Component 1")
# plt.ylabel("Principal Component 2")
# plt.colorbar()

# st.pyplot(fig)
