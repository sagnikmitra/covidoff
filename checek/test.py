from openpyxl import load_workbook
import openpyxl as pxl
import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from bokeh.models.widgets import Button
from bokeh.models import CustomJS
from streamlit_bokeh_events import streamlit_bokeh_events
from io import StringIO
from PIL import Image

gc = pd.read_excel('gc.xlsx')
exp = pxl.load_workbook('gc.xlsx')
sheet = exp['Sister Nivedita University, Kol']
st.write("""
# Check your Qwiklabs Progress for GoogleCloudFacilitator Program 2021
## Your Host: Sister Nivedita University
""")
flag = 0
milestone1_quest = 8 
milestone1_skill_badges = 4
milestone2_quest = 16
milestone2_skill_badges = 8 
milestone3_quest = 24
milestone3_skill_badges = 12 
milestone4_quest = 30
milestone4_skill_badges = 15 

selected = st.sidebar.selectbox(
    "Search By", ("Default", "Email", "Public Profile URL","Generate Your Profile Badge"))

page_name = ['Email','Public Profile URL','Generate Your Profile Badge']
page = st.radio('Search By', page_name)
image_file = st.file_uploader("Upload Image",type=['jpg','png','jpeg'])

size = (750,750)
img = Image.open("test.png").convert("RGBA")

img = img.resize(size,Image.ANTIALIAS)
card = Image.open(image_file)
# x, y = img.size
# size = (600,600)
card = card.resize(size,Image.ANTIALIAS)

card.paste(img, (0, 0), img)
st.image(card)
card.save("first.jpg", format="png")
if page == 'Email':
    str = st.text_input('Enter You Qwiklabs Email Id')
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=2).value.lower() == str.lower():
            st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
            st.write(f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
            st.write(f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
            st.write("## **Qwiklabs Public URL:**")
            st.write(sheet.cell(row=i, column=6).value)
            st.write(f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
            st.write(f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")
            
            #Quest and Skill Badges difference calculation for Milestone 1
            quest_difference_milestone_1 = (milestone1_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_1 = (milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))
            
            #Milestone 1
            if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1) :
                st.write("## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
            else:
                st.warning("You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
                if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 > 0):
                    st.write(f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                else:
                    if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 <= 0):
                        st.write(f"### You are **0 Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                    if (skill_badges_difference_milestone_1 <= 0 and quest_difference_milestone_1 > 0):
                        st.write(f"### You are **{quest_difference_milestone_1} Quests** and **0 Skill Badges** Away to Complete the First Milestone")
            
            #Quest and Skill Badges difference calculation for Milestone 2
            quest_difference_milestone_2 = (milestone2_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_2 = (milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))
            
            #Milestone 2
            if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1) :
                st.write("## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
            else:
                if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 > 0):
                    st.write(f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                else:
                    if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 <= 0):
                        st.write(f"### You are **0 Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                    if (skill_badges_difference_milestone_2 <= 0 and quest_difference_milestone_2 > 0):
                        st.write(f"### You are **{quest_difference_milestone_2} Quests** and **0 Skill Badges** Away to Complete the Second Milestone")
            
            #Quest and Skill Badges difference calculation for Milestone 3
            quest_difference_milestone_3 = (milestone3_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_3 = (milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))
            
            #Milestone 3
            if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1) :
                st.write("## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
            else:
                if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 > 0):
                    st.write(f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                else:
                    if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 <= 0):
                        st.write(f"### You are **0 Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                    if (skill_badges_difference_milestone_3 <= 0 and quest_difference_milestone_3 > 0):
                        st.write(f"### You are **{quest_difference_milestone_3} Quests** and **0 Skill Badges** Away to Complete the Third Milestone")
            
            #Quest and Skill Badges difference calculation for Milestone 4
            quest_difference_milestone_4 = (milestone4_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_4 = (milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))
            
            #Milestone 4
            if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1) :
                st.write("## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
            else:
                if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 > 0):
                    st.write(f"### You are **{quest_difference_milestone_4} Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                else:
                    if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 <= 0):
                        st.write(f"### You are **0 Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                    if (skill_badges_difference_milestone_4 <= 0 and quest_difference_milestone_4 > 0):
                        st.write(f"### You are **{quest_difference_milestone_4} Quests** and **0 Skill Badges** Away to Complete the Ultimate Milestone")
            
            flag = flag + 1

    if flag == 0:
        st.write("No Search Found for the entered Details")

if page == 'Public Profile URL':
    str = st.text_input('Enter You Qwiklabs Public URL')
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=6).value.lower() == str.lower():
            st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
            st.write(f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
            st.write(f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
            st.write("## **Qwiklabs Public URL:**")
            st.write(sheet.cell(row=i, column=6).value)
            st.write(f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
            st.write(f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")
            
            #Quest and Skill Badges difference calculation for Milestone 1
            quest_difference_milestone_1 = (milestone1_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_1 = (milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))
            
            #Milestone 1
            if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1) :
                st.write("## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
            else:
                st.warning("You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
                if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 > 0):
                    st.write(f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                else:
                    if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 <= 0):
                        st.write(f"### You are **0 Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                    if (skill_badges_difference_milestone_1 <= 0 and quest_difference_milestone_1 > 0):
                        st.write(f"### You are **{quest_difference_milestone_1} Quests** and **0 Skill Badges** Away to Complete the First Milestone")
            
            #Quest and Skill Badges difference calculation for Milestone 2
            quest_difference_milestone_2 = (milestone2_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_2 = (milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))
            
            #Milestone 2
            if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1) :
                st.write("## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
            else:
                if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 > 0):
                    st.write(f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                else:
                    if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 <= 0):
                        st.write(f"### You are **0 Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                    if (skill_badges_difference_milestone_2 <= 0 and quest_difference_milestone_2 > 0):
                        st.write(f"### You are **{quest_difference_milestone_2} Quests** and **0 Skill Badges** Away to Complete the Second Milestone")
            
            #Quest and Skill Badges difference calculation for Milestone 3
            quest_difference_milestone_3 = (milestone3_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_3 = (milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))
            
            #Milestone 3
            if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1) :
                st.write("## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
            else:
                if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 > 0):
                    st.write(f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                else:
                    if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 <= 0):
                        st.write(f"### You are **0 Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                    if (skill_badges_difference_milestone_3 <= 0 and quest_difference_milestone_3 > 0):
                        st.write(f"### You are **{quest_difference_milestone_3} Quests** and **0 Skill Badges** Away to Complete the Third Milestone")
            
            #Quest and Skill Badges difference calculation for Milestone 4
            quest_difference_milestone_4 = (milestone4_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_4 = (milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))
            
            #Milestone 4
            if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1) :
                st.write("## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
            else:
                if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 > 0):
                    st.write(f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Ultimate Milestone")
                else:
                    if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 <= 0):
                        st.write(f"### You are **0 Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                    if (skill_badges_difference_milestone_4 <= 0 and quest_difference_milestone_4 > 0):
                        st.write(f"### You are **{quest_difference_milestone_4} Quests** and **0 Skill Badges** Away to Complete the Ultimate Milestone")
            
            flag = flag + 1

    if flag == 0:
        st.write("No Search Found for the entered Details")
if page=="Generate Your Profile Badge":
    st.write("""
    # Generate your GCRF Avatar
    """)
    flag = 0
    milestone1_quest = 8 
    milestone1_skill_badges = 4
    milestone2_quest = 16
    milestone2_skill_badges = 8 
    milestone3_quest = 24
    milestone3_skill_badges = 12 
    milestone4_quest = 30
    milestone4_skill_badges = 15 
    str = st.text_input('Enter You Qwiklabs Email Id')
    image_file = st.file_uploader("Upload Image",type=['jpg','png','jpeg'])
    img1 = Image.open(r"test.jpg")
    # if image_file is not None:
    #     img1 = Image.open(image_file)
    miletry = 0
    # for i in range(2, sheet.max_row+1):
    #     if sheet.cell(row=i, column=2).value.lower() == str.lower():
    #         quest_difference_milestone_1 = (milestone1_quest - int(sheet.cell(row=i, column=7).value))
    #         skill_badges_difference_milestone_1 = (milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))
    #         quest_difference_milestone_2 = (milestone2_quest - int(sheet.cell(row=i, column=7).value))
    #         skill_badges_difference_milestone_2 = (milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))
    #         quest_difference_milestone_3 = (milestone3_quest - int(sheet.cell(row=i, column=7).value))
    #         skill_badges_difference_milestone_3 = (milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))
    #         quest_difference_milestone_4 = (milestone4_quest - int(sheet.cell(row=i, column=7).value))
    #         skill_badges_difference_milestone_4 = (milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))
            
    #         if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1) :
    #             miletry = 1
    #             break
    #         elif (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1) :
    #             miletry = 2
    #             break
    #         elif (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1) :
    #             miletry = 3
    #             break 
    #         elif (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1) :
    #             miletry = 4
    #             break  
    #         else:
    #             miletry = 0
    #             break
    # if(miletry == 1):
    #     img2 = Image.open(r"first.png")
    # if(miletry == 2):
    #     img2 = Image.open(r"second.png")
    # if(miletry == 3):
    #     img2 = Image.open(r"third.png")
    # if(miletry == 4):
    #     img2 = Image.open(r"ultimate.png")
    img2 = Image.open(r"first.png")
    img1.paste(img2, (0,0), mask = img2)
    img1.show()


# Displaying the image
# image_file.show()
st.write(" ")
st.write(" ")
st.write(" ")
st.write(" ")
st.write(" ")
st.write("#### This is a personal project and is not endorsed by Google LLC.")
st.write("#### Developed & Maintained by: **[Sagnik Mitra](https://linkedin.com/in/sagnikmitra/) & [Manish Kumar Barnwal](https://linkedin.com/in/imanishbarnwal/)** with :heart:")

