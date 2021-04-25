from openpyxl import load_workbook
import openpyxl as pxl
import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from bokeh.models.widgets import Button
from bokeh.models import CustomJS
from streamlit_bokeh_events import streamlit_bokeh_events
from io import StringIO
from PIL import Image

gc = pd.read_excel('state_city_data.xlsx')
exp = pxl.load_workbook('state_city_data.xlsx')
sheet2 = exp['Sheet2']
sheet3 = exp['Sheet3']

for i in range(2, 738):
    print("if(document.drop_list.Category.value == '{}'#".format(
        sheet3.cell(row=i, column=2).value))
    # print('addOption(document.drop_list.Category,"{}","{}",""*'.format(sheet.cell(row=i, column=1).value,
    #       sheet.cell(row=i, column=1).value))
    for j in range(2, 738):
        if(sheet3.cell(row=j, column=2).value == sheet3.cell(row=j+1, column=2).value):
            print('addOption(document.drop_list.SubCat,"{}","{}"*'.format(
                sheet3.cell(row=i, column=1).value, sheet3.cell(row=i, column=1).value))
        elif(sheet3.cell(row=j, column=2).value != sheet3.cell(row=j+1, column=2)):
            print('addOption(document.drop_list.SubCat,"{}","{}"^'.format(
                sheet3.cell(row=i, column=1).value, sheet3.cell(row=i, column=1).value))
            break
    # print('addOption(document.drop_list.Category,"{}","{}",""*'.format(sheet.cell(row=i, column=1).value,
    # print("^")
    #       sheet.cell(row=i, column=1).value))

'''
if (document.drop_list.Category.value == 'Fruits') {
        addOption(document.drop_list.SubCat, "Mango", "Mango");
        addOption(document.drop_list.SubCat, "Banana", "Banana");
        addOption(document.drop_list.SubCat, "Orange", "Orange");
    }
'''

st.write(" ")
st.write(" ")
st.write(" ")
st.write(" ")
st.write("[Check your Qwiklabs Syllabus here](https://bitly/crf-syllabus)")
st.write("Deadline of the Program: 10th June 11.30PM ")
st.write("#### This is a personal project and is not endorsed by Google LLC.")
st.write(
    "#### Developed & Maintained by: **[Sagnik Mitra](https://linkedin.com/in/sagnikmitra/) & [Manish Kumar Barnwal](https://linkedin.com/in/imanishbarnwal/)** with :snake: & :heart:")
# st.video('https://www.youtube.com/watch?v=Lf_tQWluHWA&t=10s')
