import streamlit as st
import numpy as np
import pandas as pd
import sklearn
from sklearn import datasets
from sklearn.neighbors import KNeighborsClassifier
from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
from sklearn.decomposition import PCA
import matplotlib.pyplot as plt


gc = pd.read_excel('gc.xlsx')
import openpyxl as pxl
from openpyxl import load_workbook

exp = pxl.load_workbook('gc.xlsx')
sheet = exp['Sister Nivedita University, Kol']

str = input("Enter Email: ")
for i in range(2, sheet.max_row+1):
    if sheet.cell(row=i, column=2).value == str:
        print(sheet.cell(row=i, column=5).value)
        print(sheet.cell(row=i, column=6).value)
        print(sheet.cell(row=i, column=7).value)
        print(sheet.cell(row=i, column=8).value)
# st.write("""
# # Check your Qwiklabs Progress
# """)
# dataset_name = st.sidebar.selectbox("Select Dataset",("Iris","Breast Cancer","Wine Dataset"))
# st.write(f"## Name of the Dataset: {dataset_name}")
# classifier_name = st.sidebar.selectbox("Select Classifier",("KNN","SVM","Random Forest"))