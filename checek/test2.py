from openpyxl import load_workbook
import openpyxl as pxl
import streamlit as st
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from bokeh.models.widgets import Button
from bokeh.models import CustomJS
from streamlit_bokeh_events import streamlit_bokeh_events
from io import StringIO
from PIL import Image

gc = pd.read_excel('gc.xlsx')
exp = pxl.load_workbook('gc.xlsx')
sheet = exp['Sister Nivedita University, Kol']


st.write("""
# Check your Qwiklabs Progress for GoogleCloudReady Facilitator Program
## Your Host: Sister Nivedita University
""")
st.image('crf21-rect-banner.png')
flag = 0
milestone1_quest = 8
milestone1_skill_badges = 4
milestone2_quest = 16
milestone2_skill_badges = 8
milestone3_quest = 24
milestone3_skill_badges = 12
milestone4_quest = 30
milestone4_skill_badges = 15

selected = st.sidebar.selectbox(
    "Search By", ("Default", "Email", "Public Profile URL", "Generate Your Profile Badge", "Milestone Achievers", "Syllabus"))

page_name = ['Check Your Progress by Email Id',
             'Check Your Progress by Qwiklabs Public URL', 'Generate Your Profile Badge', 'Milestone Achievers']
page = st.radio('What do you want to do today?', page_name)

if page == 'Check Your Progress by Email Id':
    str = st.text_input('Enter You Qwiklabs Email Id')
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=2).value.lower() == str.lower():
            st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
            st.write(
                f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
            st.write(
                f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
            st.write("## **Qwiklabs Public URL:**")
            st.write(sheet.cell(row=i, column=6).value)
            st.write(
                f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
            st.write(
                f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")

            # Quest and Skill Badges difference calculation for Milestone 1
            quest_difference_milestone_1 = (
                milestone1_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_1 = (
                milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))

            # Milestone 1
            if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1):
                st.write(
                    "## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
            else:
                st.warning(
                    "You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
                if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                else:
                    if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                    if (skill_badges_difference_milestone_1 <= 0 and quest_difference_milestone_1 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_1} Quests** and **0 Skill Badges** Away to Complete the First Milestone")

            # Quest and Skill Badges difference calculation for Milestone 2
            quest_difference_milestone_2 = (
                milestone2_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_2 = (
                milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))

            # Milestone 2
            if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
            else:
                if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                else:
                    if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                    if (skill_badges_difference_milestone_2 <= 0 and quest_difference_milestone_2 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_2} Quests** and **0 Skill Badges** Away to Complete the Second Milestone")

            # Quest and Skill Badges difference calculation for Milestone 3
            quest_difference_milestone_3 = (
                milestone3_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_3 = (
                milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))

            # Milestone 3
            if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
            else:
                if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                else:
                    if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                    if (skill_badges_difference_milestone_3 <= 0 and quest_difference_milestone_3 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_3} Quests** and **0 Skill Badges** Away to Complete the Third Milestone")

            # Quest and Skill Badges difference calculation for Milestone 4
            quest_difference_milestone_4 = (
                milestone4_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_4 = (
                milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))

            # Milestone 4
            if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
            else:
                if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_4} Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                else:
                    if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                    if (skill_badges_difference_milestone_4 <= 0 and quest_difference_milestone_4 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_4} Quests** and **0 Skill Badges** Away to Complete the Ultimate Milestone")

            flag = flag + 1

    if flag == 0:
        st.write("No Search Found for the entered Details")

if page == 'Check Your Progress by Qwiklabs Public URL':
    str = st.text_input('Enter You Qwiklabs Public URL')
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=6).value.lower() == str.lower():
            st.write(f"## **Name:** {sheet.cell(row=i, column=1).value}")
            st.write(
                f"## **Email Address:** {sheet.cell(row=i, column=2).value}")
            st.write(
                f"## **Enrollment Status:** {sheet.cell(row=i, column=5).value}")
            st.write("## **Qwiklabs Public URL:**")
            st.write(sheet.cell(row=i, column=6).value)
            st.write(
                f"## **No. Of Quests Completed:** {int(sheet.cell(row=i, column=7).value)}")
            st.write(
                f"## **No. Of Skill Badges Completed:** {int(sheet.cell(row=i, column=8).value)}")

            # Quest and Skill Badges difference calculation for Milestone 1
            quest_difference_milestone_1 = (
                milestone1_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_1 = (
                milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))

            # Milestone 1
            if (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1):
                st.write(
                    "## ```Congratulations! You have completed the First Milestone! On a Streak!``` :fire:")
            else:
                st.warning(
                    "You have not completed any of the Milestones. Start completing the amazing quests and skill badges to Kickstart your cloud journey and receive exciting gifts")
                if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_1} Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                else:
                    if (skill_badges_difference_milestone_1 > 0 and quest_difference_milestone_1 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_1} Skill Badges** Away to Complete the First Milestone")
                    if (skill_badges_difference_milestone_1 <= 0 and quest_difference_milestone_1 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_1} Quests** and **0 Skill Badges** Away to Complete the First Milestone")

            # Quest and Skill Badges difference calculation for Milestone 2
            quest_difference_milestone_2 = (
                milestone2_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_2 = (
                milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))

            # Milestone 2
            if (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Second Milestone! On Fire!``` :boom:")
            else:
                if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_2} Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                else:
                    if (skill_badges_difference_milestone_2 > 0 and quest_difference_milestone_2 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_2} Skill Badges** Away to Complete the Second Milestone")
                    if (skill_badges_difference_milestone_2 <= 0 and quest_difference_milestone_2 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_2} Quests** and **0 Skill Badges** Away to Complete the Second Milestone")

            # Quest and Skill Badges difference calculation for Milestone 3
            quest_difference_milestone_3 = (
                milestone3_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_3 = (
                milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))

            # Milestone 3
            if (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Third Milestone! Unstoppable!``` :star2:")
            else:
                if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                else:
                    if (skill_badges_difference_milestone_3 > 0 and quest_difference_milestone_3 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Third Milestone")
                    if (skill_badges_difference_milestone_3 <= 0 and quest_difference_milestone_3 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_3} Quests** and **0 Skill Badges** Away to Complete the Third Milestone")

            # Quest and Skill Badges difference calculation for Milestone 4
            quest_difference_milestone_4 = (
                milestone4_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_4 = (
                milestone4_skill_badges-int(sheet.cell(row=i, column=8).value))

            # Milestone 4
            if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1):
                st.write(
                    "## ```Congratulations! You have completed the Ultimate Milestone!! True Legend!!``` :heart_eyes:")
            else:
                if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 > 0):
                    st.write(
                        f"### You are **{quest_difference_milestone_3} Quests** and **{skill_badges_difference_milestone_3} Skill Badges** Away to Complete the Ultimate Milestone")
                else:
                    if (skill_badges_difference_milestone_4 > 0 and quest_difference_milestone_4 <= 0):
                        st.write(
                            f"### You are **0 Quests** and **{skill_badges_difference_milestone_4} Skill Badges** Away to Complete the Ultimate Milestone")
                    if (skill_badges_difference_milestone_4 <= 0 and quest_difference_milestone_4 > 0):
                        st.write(
                            f"### You are **{quest_difference_milestone_4} Quests** and **0 Skill Badges** Away to Complete the Ultimate Milestone")

            flag = flag + 1

    if flag == 0:
        st.write("No Search Found for the entered Details")
if page == "Generate Your Profile Badge":
    st.write("""
    ## **#GoogleCloudReady** Facilitator Program Badge
    """)
    flag = 0
    milestone1_quest = 8
    milestone1_skill_badges = 4
    milestone2_quest = 16
    milestone2_skill_badges = 8
    milestone3_quest = 24
    milestone3_skill_badges = 12
    milestone4_quest = 30
    milestone4_skill_badges = 15
    miletry = 0
    str = st.text_input('Enter You Qwiklabs Email Id')
    for i in range(2, sheet.max_row+1):
        if sheet.cell(row=i, column=2).value.lower() == str.lower():
            quest_difference_milestone_1 = (
                milestone1_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_1 = (
                milestone1_skill_badges-int(sheet.cell(row=i, column=8).value))
            quest_difference_milestone_2 = (
                milestone2_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_2 = (
                milestone2_skill_badges-int(sheet.cell(row=i, column=8).value))
            quest_difference_milestone_3 = (
                milestone3_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_3 = (
                milestone3_skill_badges-int(sheet.cell(row=i, column=8).value))
            quest_difference_milestone_4 = (
                milestone4_quest - int(sheet.cell(row=i, column=7).value))
            skill_badges_difference_milestone_4 = (
                milestone4_skill_badges - int(sheet.cell(row=i, column=8).value))

            if (quest_difference_milestone_4 < 1) and (skill_badges_difference_milestone_4 < 1):
                miletry = 4
                break
            elif (quest_difference_milestone_3 < 1) and (skill_badges_difference_milestone_3 < 1):
                miletry = 3
                break
            elif (quest_difference_milestone_2 < 1) and (skill_badges_difference_milestone_2 < 1):
                miletry = 2
                break
            elif (quest_difference_milestone_1 < 1) and (skill_badges_difference_milestone_1 < 1):
                miletry = 1
                break
            else:
                miletry = 0
                break

    st.success(f"You're Currently on Milestone {miletry}")
    st.write(
        "### **Instructions on Uploading your Image and Downloading the Badge:**")
    st.write(f"""
    * Click on Browse Files below to Upload an image
    * Upload a Square Image to get the best version of your Badge
    * If you upload a landscape or out of shape image, it would be resized to 1:1
    * According to your Milestone, your picture will be automatically applied with a badge
    * Right click on the Image and select save image as to Download the file
    * Then do share your badge on [Discord](https://bit.ly/crf-discord) and on your social media handles by tagging us as your Facilitator and Google Cloud India, also use `#GoogleCloudReady` tag. Google Cloud team closely monitor this tag :smile: :tada:
    * Find the Video Tutorial Below:
    """)
    st.write(" ")
    st.video('https://www.youtube.com/watch?v=8m9hTYNPaK8')
    image_file = st.file_uploader("Upload Image", type=['jpg', 'png', 'jpeg'])
    if image_file is not None:
        size = (750, 750)
        if miletry == 1:
            img = Image.open("first.png").convert("RGBA")
        elif miletry == 2:
            img = Image.open("second.png").convert("RGBA")
        elif miletry == 3:
            img = Image.open("third.png").convert("RGBA")
        elif miletry == 4:
            img = Image.open("ultimate.png").convert("RGBA")
        elif miletry == 0:
            img = Image.open("nomile.png").convert("RGBA")
        img = img.resize(size, Image.ANTIALIAS)
        card = Image.open(image_file)

        card = card.resize(size, Image.ANTIALIAS)

        card.paste(img, (0, 0), img)
        card.save("first.jpg", format="png")
        st.image(card)
    miletry = 0
if page == "Milestone Achievers":
    list1 = []
    list2 = []
    list3 = []
    list4 = []
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=7).value >= milestone4_quest and sheet.cell(row=i, column=8).value >= milestone4_skill_badges:
            list4.append(sheet.cell(row=i, column=1).value.title())
        if sheet.cell(row=i, column=7).value >= milestone3_quest and sheet.cell(row=i, column=8).value >= milestone3_skill_badges:
            list3.append(sheet.cell(row=i, column=1).value.title())
        if sheet.cell(row=i, column=7).value >= milestone2_quest and sheet.cell(row=i, column=8).value >= milestone2_skill_badges:
            list2.append(sheet.cell(row=i, column=1).value.title())
        if sheet.cell(row=i, column=7).value >= milestone1_quest and sheet.cell(row=i, column=8).value >= milestone1_skill_badges:
            list1.append(sheet.cell(row=i, column=1).value.title())
    list1.sort()
    list2.sort()
    list3.sort()
    list4.sort()
    st.write("""
    ### **Achievers of Ultimate Milestone :star:**
    """)
    listindex = 0
    if (len(list4) == 0):
        st.info("No one reached the Ultimate Milestone yet.")
    else:
        for i in range(0, len(list4)):
            st.write(f"**{listindex+1}**: {list4[i]}")
            listindex += 1
    listindex = 0
    st.write("""
    ### **Achievers of Third Milestone :three:**
    """)
    for i in range(0, len(list3)):
        if list3[i] not in list4:
            st.write(f"**{listindex+1}**: {list3[i]}")
            listindex += 1
    listindex = 0
    st.write("""
    ### **Achievers of Second Milestone :two:**
    """)
    for i in range(0, len(list2)):
        if list2[i] not in list3:
            st.write(f"**{listindex+1}**: {list2[i]}")
            listindex += 1
    listindex = 0
    st.write("""
    ### **Achievers of First Milestone :one:**
    """)
    for i in range(0, len(list1)):
        if list1[i] not in list2:
            st.write(f"**{listindex+1}**: {list1[i]}")
            listindex += 1
    st.image('gift.png')

st.write(" ")
st.write(" ")
st.write(" ")
st.write(" ")
st.write("[Check your Qwiklabs Syllabus here](https://bitly/crf-syllabus)")
st.write("Deadline of the Program: 10th June 11.30PM ")
st.write("#### This is a personal project and is not endorsed by Google LLC.")
st.write(
    "#### Developed & Maintained by: **[Sagnik Mitra](https://linkedin.com/in/sagnikmitra/) & [Manish Kumar Barnwal](https://linkedin.com/in/imanishbarnwal/)** with :snake: & :heart:")
# st.video('https://www.youtube.com/watch?v=Lf_tQWluHWA&t=10s')
